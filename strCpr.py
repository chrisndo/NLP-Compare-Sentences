import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
from difflib import SequenceMatcher
from fuzzywuzzy import fuzz, process
import jellyfish as jf
import nltk, string
from nltk.corpus import stopwords
from nltk.corpus import wordnet as wn
from nltk.stem import WordNetLemmatizer
from sklearn.feature_extraction.text import TfidfVectorizer
import ssl
import sys, argparse
import warnings
import time
import spacy
import wmd


'''
#################### Notes ####################
https://stackoverflow.com/questions/15173225/calculate-cosine-similarity-given-2-sentence-strings
#https://dev.to/coderasha/compare-documents-similarity-using-python-nlp-4odp
#https://stackoverflow.com/questions/8897593/how-to-compute-the-similarity-between-two-text-documents
#https://spacy.io/usage/vectors-similarity
#http://www.cs.cornell.edu/~kilian/papers/wmd_metric.pdf

#sample inputs
A='Monitor remote connections, detect aberrant behavior and notify appropriate personnel.'
B='Aberrant behavior on remote connections must be detected. Appropriate personnel must be notified as appropriate.'
C='Temporary access to cloud environment resources shall be restricted, monitored, and revoked within 24 hours of expiration.'
D='Temporary and or conditional access to cloud environment resources must be restricted based on least privileged, monitored, and must be revoked within organizational defined period of expiration'

#to do
-look at whoosh
-look into synonyms

#command line string to run script
python3 strCpr.py --file1 input.xlsx --file2 input.xlsx --sheet1 US --sheet2 Global --col1 Control --col2 'Requirement description'
python3 strCpr.py --file1 input_sm.xlsx --file2 input_sm.xlsx --sheet1 US --sheet2 Global --col1 Control --col2 'Requirement description'
file1 = 'input.xlsx'
file2 = 'input.xlsx'
sheet1 = 'US'
sheet2 = 'Global'
column1 = 'Control'
column2 = 'Requirement description'
'''


#################### Functions & global vars ####################
remove_punctuation_map = dict((ord(char), None) for char in string.punctuation)

#SequenceMatcher
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

#begin NLTK functions
def stem_tokens(tokens): #for NLTK Cosine Analysis
    return [stemmer.stem(item) for item in tokens]

def lemma_tokens(tokens): #for use with lemma normalization (not being used, but may be better)
    return [wnl.lemmatize(item) for item in tokens]

def stem_and_synonym(tokens):
    stemmed=[]
    orig_items = []
    for item in tokens:
        orig_items.append(item)
        stemmed.extend(get_synonyms(item))
        stemmed.append(stemmer.stem(item))
    return stemmed

def get_synonyms(word):
    synonyms = []
    for syns in wn.synsets(word):
        for item in syns.lemma_names():
            if item.lower() not in synonyms:
                synonyms.append(item.lower())
    return synonyms

#remove punctuation, lowercase, stem
def normalize(text): #for NLTK Cosine Analysis
    return stem_tokens(nltk.word_tokenize(text.lower().translate(remove_punctuation_map)))

def normalize_lemma(text): #for use with lemma normalization (not being used, but may be better)
    return lemma_tokens(nltk.word_tokenize(text.lower().translate(remove_punctuation_map)))

def normalize_syn(text): #for use with lemma normalization (not being used, but may be better)
    return stem_and_synonym(nltk.word_tokenize(text.lower().translate(remove_punctuation_map)))

def cosine_sim(text1, text2): #for NLTK Cosine Analysis
    tfidf = vectorizer.fit_transform([text1, text2]) #use vectorizer instead of vectorizer2 for nltk stopword list
    return ((tfidf * tfidf.T).A)[0,1]

def cosine_sim_syn(text1, text2): #for NLTK Cosine Analysis
    tfidf = vectorizer_syn.fit_transform([text1, text2]) #use vectorizer instead of vectorizer2 for nltk stopword list
    return ((tfidf * tfidf.T).A)[0,1]

#end NLTK Functions


#################### Might be needed to download files with SSL ####################
'''
try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context
'''
#nltk.download('punkt') # if necessary...
#nltk.download('stopwords') #if necessary

class init_args(object):

    def __init__(self):
        self.parser = argparse.ArgumentParser()
        self.parser.add_argument(
            "--file1", type=str, help="File 1 (excel file) that contains column of search strings. " +
            "This configuration is required.", required=True)
        self.parser.add_argument(
            "--file2", type=str, help="File 2 (excel file) that contains column of search strings that will be searched against using File 1. " +
            "This configuration is required.", required=True)
        self.parser.add_argument(
            "--sheet1", type=str, help="Which sheet in File 1 is being used to search? " +
            "This configuration is required.", required=True)
        self.parser.add_argument(
            "--sheet2", type=str, help="Which sheet in File 2 is being used to search? " +
            "This configuration is required.", required=True)
        self.parser.add_argument(
            "--col1", type=str, help="Which Column in Sheet 1 / File 1 is being used to search? " +
            "This configuration is required.", required=True)
        self.parser.add_argument(
            "--col2", type=str, help="Which Column in Sheet 2 / File 2 is being used to search? " +
            "This configuration is required.", required=True)

        self.args = self.parser.parse_args(sys.argv[1:])

    def get_args(self):
        return self.args

if __name__ == "__main__":
    start_time = time.time()
    print('Starting analysis')
    args = init_args().get_args()
    file1 = args.file1
    file2 = args.file2
    sheet1 = args.sheet1
    sheet2 = args.sheet2
    column1 = args.col1
    column2 = args.col2

    #################### Initialize workbook ####################
    #file = 'input.xlsx'
    destPath = 'Results/results-'+datetime.now().strftime("%Y%m%d-%H%M%S%p")+'.xlsx'
    # Create Workbook & Sheet
    destFile = Workbook()
    destFile.save(destPath)
    book = load_workbook(destPath)
    writer = pd.ExcelWriter(destPath, engine='openpyxl')
    writer.book = book
    book.remove(book['Sheet'])
    excel_file1 = pd.ExcelFile(file1)
    excel_file2 = pd.ExcelFile(file2)
    # Convert to Dataframe
    df_f1 = pd.read_excel(excel_file1, sheet_name=sheet1, index_col=None, na_values=['NA'])
    df_f2 = pd.read_excel(excel_file2, sheet_name=sheet2, index_col=None, na_values=['NA'])
    df_combined = pd.DataFrame(df_f1[column1])

    # #################### Word Movers Distance ####################
    # compare_time = time.time()
    # lowest = 999
    # similarity = 0
    # match = ''
    # nlp = spacy.load('en_core_web_lg')
    # nlp.add_pipe(wmd.WMD.SpacySimilarityHook(nlp), last=True)
    #
    # for i,search1 in df_f1[column1].iteritems():
    #     for y,search2 in df_f2[column2].iteritems():
    #         a = nlp(search1)
    #         b = nlp(search2)
    #         similarity = a.similarity(b)
    #         if (similarity < lowest):
    #             lowest = similarity
    #             match = search2
    #     # print('\n\n%s -  \n %s - %s seconds' % (search1, search2, (time.time() - compare_time)))
    #     df_f1.at[i,'Match Percentage'] = lowest
    #     df_f1.at[i,'Closest Match'] = match
    #     lowest = 999
    #     match = ''
    #     print('row: %s - %s seconds' % (i, (time.time() - compare_time)))
    #
    # df_combined['WMD Percentage'] = df_f1['Match Percentage']
    # df_combined['WMD Match'] = df_f1['Closest Match']
    # df_f1.to_excel(writer, sheet_name='WMD Matcher', index=None)
    # df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    # print('finished Word Movers Distance Analysis - %s seconds' % (time.time() - compare_time))

    #################### SequenceMatcher ####################
    compare_time = time.time()
    highest = 0
    similarity = 0
    match = ''

    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            similarity = similar(search1, search2)
            if (highest < similarity):
                highest = similarity
                match = search2
        df_f1.at[i,'Match Percentage'] = highest
        df_f1.at[i,'Closest Match'] = match
        highest = 0
        match=''

    df_combined['SequenceMatcher Percentage'] = df_f1['Match Percentage']
    df_combined['SequenceMatcher Match'] = df_f1['Closest Match']
    df_f1.to_excel(writer, sheet_name='Sequence Matcher', index=None)
    df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    print('finished Sequence Matcher Analysis - %s seconds' % (time.time() - compare_time))

    #################### fuzzywuzzy ####################
    compare_time = time.time()
    highest = 0
    similarity = 0
    match = ''

    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            similarity = fuzz.token_set_ratio(search1, search2)
            if (highest < similarity):
                highest = similarity
                match = search2
        df_f1.at[i,'Match Percentage'] = highest
        df_f1.at[i,'Closest Match'] = match
        highest = 0
        match=''

    df_combined['FuzzyWuzzy Percentage'] = df_f1['Match Percentage']
    df_combined['FuzzyWuzzy Match'] = df_f1['Closest Match']
    df_f1.to_excel(writer, sheet_name='FuzzyWuzzy', index=None)
    df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    print('finished FuzzyWuzzy Analysis - %s seconds' % (time.time() - compare_time))

    #################### jellyfish ####################
    compare_time = time.time()
    jw_highest = jaro_highest = 0
    dl_lowest = 999
    dl_similarity = jw_similarity = jaro_similarity = 0
    dl_match = jw_match = jaro_match = ''

    #array = df_f2[column2].values

    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            dl_similarity = jf.levenshtein_distance(search1, search2)
            jw_similarity = jf.jaro_winkler(search1, search2)
            jaro_similarity = jf.jaro_distance(search1, search2)
            if (dl_lowest > dl_similarity):
                dl_lowest = dl_similarity
                dl_match = search2
            if (jw_highest < jw_similarity):
                jw_highest = jw_similarity
                jw_match = search2
            if (jaro_highest < jaro_similarity):
                jaro_highest = jaro_similarity
                jaro_match = search2
        df_f1.at[i,'DL Match Percentage'] = dl_lowest
        df_f1.at[i,'DL Closest Match'] = dl_match
        df_f1.at[i,'jw Match Percentage'] = jw_highest
        df_f1.at[i,'jw Closest Match'] = jw_match
        df_f1.at[i,'jaro Match Percentage'] = jaro_highest
        df_f1.at[i,'jaro Closest Match'] = jaro_match
        jw_highest = jaro_highest = 0
        dl_lowest = dl_lowest = 999
        dl_match = jw_match = jaro_match = ''

    df_combined['DL Match Percentage'] = df_f1['DL Match Percentage']
    df_combined['DL Closest Match'] = df_f1['DL Closest Match']
    df_combined['jw Match Percentage'] = df_f1['jw Match Percentage']
    df_combined['jw Closest Match'] = df_f1['jw Closest Match']
    df_combined['jaro Match Percentage'] = df_f1['jaro Match Percentage']
    df_combined['jaro Closest Match'] = df_f1['jaro Closest Match']
    df_f1.to_excel(writer, sheet_name='Jellyfish', index=None)
    df_f1.drop(columns=['DL Match Percentage', 'DL Closest Match','jw Match Percentage','jw Closest Match', 'jaro Match Percentage','jaro Closest Match'], inplace=True)
    print('finished JellyFish Analysis - %s seconds' % (time.time() - compare_time))

    #################### nltk cosine ####################
    #https://dev.to/coderasha/compare-documents-similarity-using-python-nlp-4odp
    #https://stackoverflow.com/questions/8897593/how-to-compute-the-similarity-between-two-text-documents
    warnings.filterwarnings("ignore", category=UserWarning)
    compare_time = time.time()

    #################### Functions for cosine comparison ####################
    stemmer = nltk.stem.porter.PorterStemmer()
    wnl = WordNetLemmatizer()
    stop_words=stopwords.words('english')
    normalized_stop_words = normalize(str(stop_words))
    vectorizer = TfidfVectorizer(tokenizer=normalize, stop_words='english')

    #################### Loop to compare arrays ####################
    highest = 0
    similarity = 0
    match = ''
    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            similarity = cosine_sim(search1, search2)
            if (highest < similarity):
                highest = similarity
                match = search2
        df_f1.at[i,'Match Percentage'] = highest
        df_f1.at[i,'Closest Match'] = match
        highest = 0
        match=''

    df_combined['Cosine Match Percentage'] = df_f1['Match Percentage']
    df_combined['Cosine Closest Match'] = df_f1['Closest Match']
    df_f1.to_excel(writer, sheet_name='Cosine Similarity', index=None)
    df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    warnings.filterwarnings("default", category=UserWarning)

    print('finished Cosine Analysis - %s seconds' % (time.time() - compare_time))


    #################### nltk cosine + synonyms ####################
    #https://dev.to/coderasha/compare-documents-similarity-using-python-nlp-4odp
    #https://stackoverflow.com/questions/8897593/how-to-compute-the-similarity-between-two-text-documents
    warnings.filterwarnings("ignore", category=UserWarning)
    compare_time = time.time()

    #################### Functions for cosine comparison ####################
    stop_words=stopwords.words('english')
    normalized_stop_words = normalize(str(stop_words))
    vectorizer2 = TfidfVectorizer(tokenizer=normalize, stop_words=normalized_stop_words)
    vectorizer_syn = TfidfVectorizer(tokenizer=normalize_syn, stop_words=normalized_stop_words)

    #################### Loop to compare arrays ####################
    highest = 0
    similarity = 0
    match = ''
    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            similarity = cosine_sim_syn(search1, search2)
            if (highest < similarity):
                highest = similarity
                match = search2
        df_f1.at[i,'Match Percentage'] = highest
        df_f1.at[i,'Closest Match'] = match
        highest = 0
        match=''

    df_combined['Cosine Synonym Match Percentage'] = df_f1['Match Percentage']
    df_combined['Cosine Synonym Closest Match'] = df_f1['Closest Match']
    df_f1.to_excel(writer, sheet_name='Cosine Synonym Similarity', index=None)
    df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    warnings.filterwarnings("default", category=UserWarning)

    print('finished Cosine Synonym Analysis - %s seconds' % (time.time() - compare_time))

    #################### Functions for spaCy comparison ####################
    # this takes a long time  (15-30 mins to complete) and the results aren't good.
    # removing for now
    #https://spacy.io/usage/vectors-similarity
    #################### Loop to compare arrays ####################
    compare_time = time.time()
    highest = 0
    similarity = 0
    match = ''

    nlp = spacy.load('en_core_web_lg')

    for i,search1 in df_f1[column1].iteritems():
        for y,search2 in df_f2[column2].iteritems():
            a = nlp(search1)
            b = nlp(search2)
            similarity = a.similarity(b)
            if (highest < similarity):
                highest = similarity
                match = search2
        df_f1.at[i,'Match Percentage'] = highest
        df_f1.at[i,'Closest Match'] = match
        highest = 0
        match=''
        #print('row: %s - %s seconds' % (i, (time.time() - compare_time)))

    df_combined['spaCy Similarity Match Percentage'] = df_f1['Match Percentage']
    df_combined['spaCy Similarity Closest Match'] = df_f1['Closest Match']
    df_f1.to_excel(writer, sheet_name='spaCy Similarity Model', index=None)
    df_f1.drop(columns=['Match Percentage','Closest Match'], inplace=True)
    print('finished spaCy Analysis - %s seconds' % (time.time() - compare_time))

    # Save
    df_combined.to_excel(writer, sheet_name='Combined Results', index=None)

    column_list = ['WMD Percentage',
                    'spaCy Similarirty Match Percentage',
                    'Cosine Match Percentage',
                    'Cosine Synonym Match Percentage',
                    'SequenceMatcher Percentage',
                    'jaro Match Percentage',
                    'jw Match Percentage',
                    'FuzzyWuzzy Percentage',
                    'DL Match Percentage',
                    'DL Closest Match']
    for col in column_list:
        if col not in df_combined.columns:
            df_combined[col] = np.nan

    newdf = pd.DataFrame(df_combined[(df_combined['Cosine Match Percentage'] > 0.5) |
            (df_combined['SequenceMatcher Percentage'] > 0.5) |
            (df_combined['jaro Match Percentage'] > 0.775) |
            (df_combined['spaCy Similarity Match Percentage'] > 0.95) |
            (df_combined['DL Match Percentage'] < 40) |
            (df_combined['jw Match Percentage'] > 0.94) |
            (df_combined['Cosine Synonym Match Percentage'] > 0.50) |
            (df_combined['WMD Percentage'] < 5) |
            (df_combined['FuzzyWuzzy Percentage'] > 70)])
    newdf.drop(columns=['Cosine Match Percentage', 'spaCy Similarity Match Percentage', 'Cosine Synonym Match Percentage','spaCy Similarirty Match Percentage', 'WMD Percentage', 'SequenceMatcher Percentage', 'DL Match Percentage',  'jaro Match Percentage', 'jw Match Percentage', 'FuzzyWuzzy Percentage'], inplace=True)
    newdf.to_excel(writer, sheet_name='High Matches (Combined)', index=None)

    writer.save()
    writer.close()

    print('Script completed analysis - %s total seconds' % (time.time() - start_time))
